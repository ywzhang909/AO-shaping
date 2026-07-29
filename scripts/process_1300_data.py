"""
Process 1300-5.xlsx: add 36×36 grid positions and lookup group/pin from wiring table.

1. Read 1300-5.xlsx (1311 rows, cols A-C with data)
2. Add 36×36 Row and Column based on col A (position number): row = (A-1)//36, col = (A-1)%36
3. Parse 1300路机柜输出线序表(1).xlsx to build a mapping:
   1300-5 row number → (组名, 引脚编号, 连接器ID)
4. Join the mapping to produce enriched output
5. Save enriched data to data/1300-5-enriched.xlsx
"""

from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pandas as pd

DOCS_DIR = Path(__file__).resolve().parent.parent / "docs" / "micro deformable mirror" / "docs"
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_1300_5 = DOCS_DIR / "1300-5.xlsx"
XLSX_WIRING = DOCS_DIR / "1300路机柜输出线序表(1).xlsx"
OUTPUT_XLSX = DATA_DIR / "1300-5-enriched.xlsx"
OUTPUT_CSV = DATA_DIR / "1300-5-enriched.csv"

GRID_SIZE = 36

# =============================================================================
# Step 1: Parse 1300-5.xlsx
# =============================================================================

def read_1300_5_data(path: Path) -> pd.DataFrame:
    """Read 1300-5.xlsx into DataFrame with columns: position, ip_group, seq."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        a, b, c, *_ = row if len(row) >= 3 else (row[0], row[1] if len(row) > 1 else None, None)
        if a is not None:
            rows.append({"row_id": i, "position": int(a), "ip_group": b, "seq": c})

    df = pd.DataFrame(rows)
    # Add 36×36 grid coordinates
    df["grid_row"] = ((df["position"] - 1) // GRID_SIZE).astype(int)
    df["grid_col"] = ((df["position"] - 1) % GRID_SIZE).astype(int)
    return df


# =============================================================================
# Step 2: Parse wiring table → mapping: 1300_row_number → (group_name, pin, connector)
# =============================================================================

def parse_wiring_table(path: Path) -> dict[int, dict]:
    """Parse 1300路机柜输出线序表(1).xlsx into a dict keyed by 1300 row number.

    Each triple (pin_header_row, row_number_row, connector_row) forms one entry.
    Structure per group (260 entries each):
      Block 1: rows N..N+2, pin 277-316→1300_row ~221-260, 40 entries
      Block 2: rows N+3..N+5, pin 222-261→1300_row ~181-220, 40 entries
      Block 3: rows N+6..N+8, pin 166-205→1300_row ~141-180, 40 entries
      Block 4: rows N+9..N+11, pin 111-150→1300_row ~101-140, 40 entries
      Block 5: rows N+12..N+14, pin 56-105→1300_row ~51-100, 50 entries
      Block 6: rows N+15..N+17, pin 1-55→1300_row ~1-50, 50 entries

    Groups:
      - 一组定义: rows 2-19 (total 260 → 1300 rows 1-260)
      - 二组定义: rows 22-39 (total 260 → 1300 rows 261-520)
      - 三组定义: rows 42-59 (total 260 → 1300 rows 521-780)
      - 四组定义: rows 62-79 (total 260 → 1300 rows 781-1040)
      - 五组定义: rows 82-99 (total 260 → 1300 rows 1041-1300)
    Returns: dict {1300_row_number: {group, pin, connector}}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    max_col = ws.max_column

    group_names = {
        2: "一组", 22: "二组", 42: "三组", 62: "四组", 82: "五组",
    }

    mapping: dict[int, dict] = {}

    for group_start, group_name in group_names.items():
        for block in range(6):
            r_pin = group_start + block * 3
            r_row = r_pin + 1
            r_conn = r_pin + 2

            for col in range(2, max_col + 1):
                v_pin = ws.cell(row=r_pin, column=col).value
                v_row = ws.cell(row=r_row, column=col).value
                v_conn = ws.cell(row=r_conn, column=col).value

                # Skip empty cells
                if v_pin is None:
                    continue
                # Skip header labels like "一组定义"
                if isinstance(v_pin, str) and "定义" in v_pin:
                    continue

                # v_row should be the 1300-5 row number (could be int or string)
                try:
                    row_num = int(str(v_row).strip())
                except (ValueError, TypeError):
                    continue

                try:
                    pin_num = int(str(v_pin).strip())
                except (ValueError, TypeError):
                    continue

                connector = str(v_conn).strip() if v_conn is not None else ""

                if row_num not in mapping:
                    mapping[row_num] = {
                        "group": group_name,
                        "pin": pin_num,
                        "connector": connector,
                    }

    return mapping


# =============================================================================
# Step 3: Merge and save
# =============================================================================

def main() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)

    print("Reading 1300-5.xlsx...")
    df = read_1300_5_data(XLSX_1300_5)
    print(f"  Read {len(df)} rows with data")

    print("Parsing wiring table...")
    wiring_map = parse_wiring_table(XLSX_WIRING)
    print(f"  Parsed {len(wiring_map)} entries from wiring table")

    # Verify wiring table covers the range
    missing_rows = set(df["row_id"]) - set(wiring_map.keys())
    if missing_rows:
        print(f"  WARNING: {len(missing_rows)} rows not found in wiring table: {sorted(missing_rows)[:10]}...")

    # Join wiring info into the dataframe
    df["组"] = df["row_id"].map(lambda r: wiring_map.get(r, {}).get("group", ""))
    df["引脚编号"] = df["row_id"].map(lambda r: wiring_map.get(r, {}).get("pin", 0))
    df["连接器"] = df["row_id"].map(lambda r: wiring_map.get(r, {}).get("connector", ""))

    # Remove row_id (internal tracking)
    df_out = df.drop(columns=["row_id"])

    # Reorder columns
    column_order = [
        "position",       # Col A: 位置序号
        "grid_row",       # 36×36行
        "grid_col",       # 36×36列
        "ip_group",       # Col B: IP组
        "seq",            # Col C: 序号
        "组",             # 组名 (一组/二组等)
        "引脚编号",       # 引脚编号 (1-330)
        "连接器",         # 连接器ID (A-B-C格式)
    ]
    df_out = df_out[column_order]

    # Rename to Chinese for better readability
    df_out.rename(columns={
        "position": "位置序号",
        "grid_row": "36×36行",
        "grid_col": "36×36列",
        "ip_group": "IP组",
        "seq": "序号",
    }, inplace=True)

    # Save
    print(f"\nSaving enriched data...")
    df_out.to_excel(OUTPUT_XLSX, index=False, engine="openpyxl")
    print(f"  XLSX: {OUTPUT_XLSX}")
    df_out.to_csv(OUTPUT_CSV, index=False)
    print(f"  CSV: {OUTPUT_CSV}")

    print(f"\nEnriched data preview (first 10 rows):")
    print(df_out.head(10).to_string())
    print(f"\nTotal records: {len(df_out)}")
    print(f"36×36 grid size: {GRID_SIZE}×{GRID_SIZE} = {GRID_SIZE*GRID_SIZE}")
    print(f"Unique IP groups: {sorted(df_out['IP组'].unique())}")
    print(f"每组 column has values: {df_out['组'].unique().tolist()}")


if __name__ == "__main__":
    main()
